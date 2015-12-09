#! /usr/bin/python3

# TODO Create Methods using openpyxl module for *.xls(x/m) support
# TODO Create abstract methods for r/w operations on Excel docs
# TODO Streamline UI
# TODO Create Unit tests (may be easier if migrated to OO design)

import argparse
import os
import codecs
import collections
import datetime
import logging
import pickle
import pdb
import subprocess
import sys
import time
import xlrd
from tkinter import *
from tkinter import ttk, messagebox, filedialog
from configparser import ConfigParser
from idlelib.ToolTip import *
from xlwt import Workbook, easyxf
from xlrd import open_workbook
from openpyxl import load_workbook
from MenuItem import MenuItem

__version__ = 'v0.12.7'

TEXT_HEADERS = MenuItem.TEXT_HEADERS
IG_FIELD_SEQUENCE = MenuItem.IG_FIELD_SEQUENCE
INTEGER_FIELDS = MenuItem.INTEGER_FIELDS
STRING_FIELDS = MenuItem.STRING_FIELDS
PRICE_ARRAY_REGEX = re.compile(r'(?<=\{)[^(\{|\})].+?(?=\})')
QUOTED_COMMAS_REGEX = re.compile(r'((?<=")[^",\{\}]+),([^"\{\}]*(?="))')
BARCODE_REGEX = re.compile(r'(?<=")\d+(?=")')
PRICE_COLUMN = 8
MAX_STRING_LENGTH = 38
file_type_filters = [('Supported Files', '.xls .xlsx .txt'),
                     ('Text Files', '.txt'),
                     ('Excel Files', '.xls .xlsx .csv'), ('All Files', '.*')]
APP_DIR = os.path.join(os.getenv('APPDATA'), 'Agilysys Format Tools')
CONFIG_FILE = os.path.join(APP_DIR, 'config.ini')
LOG_FILE = os.path.join(APP_DIR, 'errors.log')
config = ConfigParser()
config.read(CONFIG_FILE)

parser = argparse.ArgumentParser()
parser.add_argument("--log", help="set logging level")
args = parser.parse_args()
if args.log:
    log_level = getattr(logging, args.log.upper(), None)
try:
    if not isinstance(log_level, int):
        log_level = logging.WARNING
except NameError:
    log_level = logging.WARNING
log_formatter = logging.Formatter(fmt='%(asctime)s %(message)s',
                                  datefmt='%H:%M:%S | ')
root_logger = logging.getLogger()
file_handler = logging.FileHandler(LOG_FILE)
root_logger.addHandler(file_handler)
root_logger.setLevel(log_level)

sep = '-' * 70
logging.error('\n\n{0}\n{1}'
              .format(time.strftime('%a %m/%d/%y %I:%M:%S%p'), sep))

console_handler = logging.StreamHandler()
root_logger.addHandler(console_handler)
file_handler.setFormatter(log_formatter)
console_handler.setFormatter(log_formatter)
if not getattr(sys, 'frozen', False):
    logging.error('Log Level={0}'.format(logging.getLevelName(log_level)))
    logging.error('Python {}'.format(sys.version))

IG_EXPORT = 1
EXCEL_FILE = 3
itemList = []
itemMap = {}


def open_file(options=None):
    """Generates File Open dialog and alters UI based on file selected."""
    logging.debug('Launching Open File Dialog')
    hide_all_buttons()
    init_dir = ''
    global checkbox_variable_map, all_boxes_selected
    all_boxes_selected = False
    checkbox_variable_map = {}
    if itemList:
        itemList.clear()

    try:
        init_dir = config['Paths']['last dir']
    except KeyError:
        init_dir = os.path.expanduser('~')

    if not options:
        options = dict()
        options['defaultextension'] = '.xls*, .txt'
        options['filetypes'] = file_type_filters
        options['title'] = 'Open...'
        options['initialdir'] = init_dir
    file_opt = options
    global in_file
    in_file = filedialog.askopenfilename(**file_opt)
    file_display_string.set(directory_display(in_file))
    options = None
    if not in_file or in_file == "":
        logging.warning("No file selected")
        return

    try:
        if get_file_type(in_file) == IG_EXPORT:
            logging.debug('IG_EXPORT found')
            pre_parse_ig_file(in_file)
            if 'duplicate' not in in_file and 'barcode' not in in_file:
                logging.debug('preparing to test for duplicate barcodes')
                file_parts = str(os.path.basename(in_file)) \
                    .rsplit('.', maxsplit=1)
                barcodes_filename = '{}_duplicate_barcodes.txt' \
                    .format(file_parts[0])
                dupe_barcodes_file = os.path.join(os.path.dirname(in_file),
                                                  barcodes_filename)
                duplicate_sku_test(dupe_barcodes_file)
            for button in simplifyButtons:
                show_button(button)
        else:
            logging.debug('EXCEL found')
            show_button(button_ig)

        config['Paths'] = {'last dir': in_file}
        with open(CONFIG_FILE, 'w') as f:
            config.write(f)

    except IOError:
        messagebox.showinfo(title='Oops',
                            message='This file is not supported.')
        logging.error('{0}\n{1}'.format(sys.exc_info()[0], sys.exc_info()[1]))
        return


def directory_display(directory):
    """Returns truncated directory string when length over 38 chars."""
    if len(directory) > MAX_STRING_LENGTH:
        split_dirs = directory.split('/')
        directory = '/'.join([split_dirs[0], '...',
                              split_dirs[-2], split_dirs[-1]])
        if len(directory) > MAX_STRING_LENGTH:
            return split_dirs[-1]
    return directory


def write_to_text_file(file, *args):
    """Writes contents of one or more lists to a file. One item per line."""
    output_list = []
    logging.debug('Preparing to write data to {}'.format(file))
    with open(file, 'w+') as f:
        for arg in args:
            f.write('\n'.join(arg))


def save_file_as(options):
    """Generates Save As dialog and returns chosen file path."""
    file_opt = options
    file_save_path = filedialog.asksaveasfilename(**file_opt)
    if file_save_path is None or file_save_path == "":
        logging.info("No file selected")
    return file_save_path


def replace_commas(match, ):
    """returns string with semi-colon substitutes for matched commas."""
    match = str(match.group(0))
    return match.replace(",", ";")


def pre_parse_ig_file(file_name):
    """Parses lines from IG export into workable objects."""
    with codecs.open(file_name, 'r', 'latin-1') as export:
        logging.debug('pre-parse initiated')
        for line in export:
            itemDetails = re.sub(PRICE_ARRAY_REGEX, replace_commas, line)
            itemDetails = re.sub(QUOTED_COMMAS_REGEX, replace_commas,
                                 itemDetails)
            item = itemDetails.split(",")
            try:
                i = MenuItem(
                    item[1], item[2], item[3], item[4], item[5],
                    item[6], item[7], item[8], item[9], item[10],
                    item[11], item[12], item[13], item[14], item[15],
                    item[16], item[18], item[19], item[20], item[21],
                    item[22], item[23], item[24], item[25], item[26],
                    item[28], item[29], item[30], item[31]
                )
            except IndexError:
                response = None
                if item[1]:
                    response = messagebox.askokcancel(
                        title='Error reading file',
                        message='Unable to parse line {0}'.format(item[1]))
                else:
                    response = messagebox.askokcancel(
                        title='Error reading file',
                        message='Unable to parse info:\n' + line)
                if not response:
                    os._exit(1)
            # Skip Store Items
            if str(i.store_id) == '0':
                itemList.append(i)
            else:
                logging.debug('skipping item {}'.format(i))
                continue
    logging.info("parse completed")


# Might be worth moving this to MenuItem class
def count_price_levels():
    """Returns total number of price levels"""
    logging.debug('counting price levels')
    num_price_levels = 0
    price_level_list = []
    for item in itemList:
        levels = item.get_prices_dict()
        for level in levels.keys():
            if level not in price_level_list:
                price_level_list.append(level)
                # if max(k for k in levels.keys()) > num_price_levels:
                #     num_price_levels = max(k for k in levels.keys())
    logging.debug('found {} price levels'.format(num_price_levels))
    return num_price_levels, price_level_list


# noinspection PyShadowingNames,PyShadowingNames
def generate_custom_excel_spreadsheet(excel_file, items=None):
    """Generates Excel spreadsheet from IG Export file

    Keyword arguements:
    excel_file -- file path to save resulting workbook
    items -- list of items to parse
    """
    items = items or itemList
    logging.debug('preparing to convert to custom Excel')
    book = Workbook()
    heading = easyxf(
        'font: bold True;'
        'alignment: horizontal center;'
    )
    oopsStyle = (easyxf('pattern: pattern solid, fore_color rose'))
    sheet = book.add_sheet('Sheet 1')
    sheet.panes_frozen = True
    sheet.remove_splits = True
    sheet.horz_split_pos = 1
    heading_row = sheet.row(0)
    keyname_row = sheet.row(1)
    headers = ['"A"', 'ID']
    keynames = ['modType', 'id']
    pricePos = 100

    for k, v in sorted(checkbox_variable_map.items(),
                       key=lambda x: IG_FIELD_SEQUENCE.get(x[0])):
        if str(v.get()) == '1':
            headers.append(TEXT_HEADERS[k])
            keynames.append(k)

    logging.debug('headers created')
    if 'Prices' in headers:
        logging.debug('Parsing header prices')
        pricePos = headers.index('Prices')
        del headers[pricePos]
        del keynames[pricePos]
        logging.debug('preparing to count price levels')
        num_price_levels, price_level_list = count_price_levels()
        price_headers = []

        # Need additional logic to filter empty price levels
        if price_level_list:
            logging.debug('adding prices from price level list')
            for price_level in sorted(price_level_list):
                price_headers.append('Price Level ' + str(price_level))
        else:
            logging.debug('adding prices from num price levels')
            for price_level in range(num_price_levels):
                price_headers.append('Price Level ' + str(price_level + 1))

        for level in reversed(price_headers):
            headers.insert(pricePos, level)
            keynames.insert(pricePos, ('priceLvl{0}'.format(level[12:])))

    # Write Headers
    for header, key, row in zip(headers, keynames, range(len(headers))):
        heading_row.write(row, header, heading)
        keyname_row.write(row, key, heading)

    logging.debug('headers written')
    # Hiding keyname row
    sheet.row(1).hidden = True

    # Write Rows
    for r, item in zip(range(2, len(items) + 2), items):
        logging.debug('writing row {0}'.format(r))
        global row_is_misaligned
        row_is_misaligned = False

        row = sheet.row(r)

        # Write item values to columns
        for col, key in enumerate(keynames):
            if 'priceLvl' in key:
                # Extract number from priceLvl key
                p = int(key[key.find('l') + 1:])
                if p in item.get_prices_dict():
                    price = item.get_prices_dict()[p]
                else:
                    price = ''

                row.write(col, str(price))
            else:
                if key in INTEGER_FIELDS:
                    row.write(col, safeIntCast(item.__dict__[key]))
                elif key == 'modType':
                    continue
                else:
                    row.write(col, str(item.__dict__[key]))

        if row_is_misaligned:
            row.write(0, 'X', oopsStyle)

    try:
        book.save(excel_file)
        messagebox.showinfo(title='Success',
                            message='Custom Excel export created successfully')
    except PermissionError:
        messagebox.showerror(title='Error', message='Unable to save file')


def generateIGUpdate(excel_file, ig_text_file):
    """Legacy function for generating IG import from full Excel workbook."""
    logging.debug('preparing to generate IG Update file using legacy function')

    sheet = book.sheet_by_index(0)
    includeColumns = set()

    for col in range(3, sheet.ncols):
        if sheet.cell_value(1, col):
            includeColumns.add(col)

    includeColumns = sorted(includeColumns)

    for row in range(2, sheet.nrows):
        itemProperties = []
        updateType = sheet.cell_value(row, 1)
        if updateType != 'A' and updateType != 'U' and \
                        updateType != 'D' and updateType != 'X':
            continue
        elif updateType == 'X':
            messagebox.showwarning(
                title='File Error',
                message='One or more lines are not aligned properly.'
                        '\nPlease correct and retry.')
            return
        else:
            itemProperties.append('"' + str(sheet.cell_value(row, 1)) + '"')
        itemProperties.append(safeIntCast((sheet.cell_value(row, 2))))
        previousIndex = 2
        for col in includeColumns:
            emptySpaces = col - previousIndex - 1
            for _ in range(emptySpaces):
                itemProperties.append('')
            if col in STRING_FIELDS:
                itemProperties.append(
                    '"' + str(sheet.cell_value(row, col)) + '"')
            else:
                itemProperties.append(safeIntCast(sheet.cell_value(row, col)))
            previousIndex = col
        if len(itemProperties) < 32:
            for _ in range(32 - len(itemProperties)):
                itemProperties.append('')
        line = ",".join(itemProperties).replace(";", ",")
        ig_text_file.write(line + "\r\n")

    messagebox.showinfo(title='Success',
                        message='IG Item Import created successfully.')


# TODO rewrite to leverage MenuItem class
def generate_ig_import(book, ig_text_file):
    """Generates IG Import File from custom Excel workbook.

    Keyword arguments:
    book -- Excel workbook (custom)
    ig_text_file -- text file to be generated for Agilysys
    """
    logging.debug('Generating IG Import from custom Excel')
    sheet = book.sheet_by_index(0)
    updated_items = False
    valid_update_types = ['A', 'U', 'D', 'X']

    for row in range(2, sheet.nrows):
        item_properties = []
        item_property_map = dict()
        price_level_map = dict()
        update_type = sheet.cell_value(row, 0)
        if update_type not in valid_update_types:
            continue
        elif update_type == 'X':
            messagebox.showwarning(
                title='File Error',
                message='One or more lines are not aligned properly.'
                        '\nPlease correct and retry.')
            return
        else:
            item_properties.append('"{0}"'.format(update_type))
            updated_items = True

        for col in range(1, sheet.ncols):
            key = sheet.cell_value(1, col)
            if 'priceLvl' in key:
                price_level_map[key] = sheet.cell_value(row, col)
            item_property_map[key] = sheet.cell_value(row, col)

        if price_level_map:
            item_property_map['price_levels'] = \
                build_ig_price_array(price_level_map)

        for key, field in sorted(IG_FIELD_SEQUENCE.items(),
                                 key=lambda x: x[1]):
            if key in item_property_map.keys():
                if field in STRING_FIELDS:
                    item_properties.append('"{0}"'.format(
                        item_property_map[key].strip('\r\n')))
                else:
                    item_properties.append(safeIntCast(item_property_map[key]))
            else:
                item_properties.append('')

        line = ','.join(item_properties).replace(';', ',')
        ig_text_file.write(line + '\r\n')

    # adding sentinel item
    ig_text_file.write(
        '"A",7110001,"{0}",,,,{{1,$0.00}},,,,,,,,,,,,,,,,,,,,,,,,,'.format(
            time.strftime('%c', time.localtime())))
    if updated_items:
        logging.info('IG Item Import created successfully')
        messagebox.showinfo(title='Success',
                            message='IG Item Import created successfully.')
    else:
        logging.info('Empty IG Item Import created')
        messagebox.showinfo(
            title='Oops',
            message="No items processed.  "
                    "Did you remember to put a 'U' or 'A'"
                    " in the first column?")


def get_items_from_xlsx(book):
    """Returns item details from xlsx book"""
    sheet = book['Menu Items & Pricing']
    item_list = []

    for i, row in enumerate(sheet.rows):
        if i <= 5:
            continue
        fields = []
        price_map = dict()

        for j, cell in enumerate(row):
            if j < 8:
                property = cell.value
                fields.append(property)
            else:
                price_level = j
                level = price_level - 7
                price = row[price_level].value or ''
                if price == '':
                    continue
                price_map[level] = price
        if price_map:
            fields.append(price_map)
        else:
            fields.append(None)

        item_list.append(fields)
    return item_list


def get_items_from_old_excel(book):
    """Returns item details from xls book"""
    sheet = book.sheet_by_name('Menu Items & Pricing')
    item_list = []

    for row in range(5, sheet.nrows):
        fields = []
        price_map = dict()

        for col in range(1, 8):
            property = sheet.cell_value(row, col)
            fields.append(property)
        for price_level in range(8, sheet.ncols):
            level = price_level - 7
            if sheet.cell_value(row, price_level) == '':
                continue
            else:
                price = sheet.cell_value(row, price_level)
            price_map[level] = price
        if price_map:
            fields.append(price_map)
        else:
            fields.append(None)
        item_list.append(fields)
    return item_list


def get_max_item_id(book):
    """Returns maximum item id from ig worksheet in book"""
    try:
        ig_sheet = book.sheet_by_name('InfoGenesis')
        max_item_id = int(ig_sheet.cell_value(ig_sheet.nrows - 1, 1))
    except AttributeError:
        ig_sheet = book['InfoGenesis']
        max_item_id = int(list(ig_sheet.rows)[-1][1].value)
    return max_item_id


# TODO Correct/Replace save path selection
def generate_standardized_ig_imports(book, save_path):
    """Generates IG import files from POS Configuration Worksheet.

    Keyword arguments:
    book -- Excel workbook
    save_path -- filename used as base for priced and unpriced output files
    """
    logging.debug('Generating standardized IG Import files')
    product_classes = get_product_classes(book)
    revenue_categories = get_revenue_categories(book)
    update_type = '"A"'
    item_list = []

    ig_priced_file = '{0} - MI_Imp_priced.txt'.format(save_path)
    ig_unpriced_file = '{0} - MI_Imp_unpriced.txt'.format(save_path)

    max_item_id = get_max_item_id(book)
    last_item_id = 0
    item_ids = set()

    revenue_category_errors = []
    product_class_errors = []
    priced_items = []
    unpriced_items = []
    if isinstance(book, xlrd.book.Book):
        items = get_items_from_old_excel(book)
    else:
        items = get_items_from_xlsx(book)

    for i, item in enumerate(items, start=1):
        use_weight = 0
        prices = '{1,$0.00}'
        fields = item[:8]
        price_map = item[8]

        if price_map:
            for k, p in price_map.items():
                try:
                    float(p)
                except ValueError:
                    if 'lb' in p:
                        price = re.split('[^0-9.]+', p)[0]
                        price_map[k] = price
                        use_weight = 1
                    else:
                        logging.warning('Skipping item {0} because price was wrong'
                                        .format(i))
                        continue

            prices = build_ig_price_array(price_map)

        rev_cat = fields[6]
        name = fields[4]
        prod_class = fields[7]
        sku = fields[3] or ''

        # Skip junk items
        if type(name) is not str:
            logging.warning('Skipping item {0} because name is invalid'.format(i))
            continue

        try:
            revenue_category = revenue_categories[rev_cat]
        except KeyError:
            if rev_cat not in revenue_category_errors:
                revenue_category_errors.append(rev_cat)
                logging.warning('Unable to get value for {0} category'.format(rev_cat))
            continue

        try:
            product_class = product_classes[prod_class]
        except KeyError:
            if prod_class not in product_class_errors:
                product_class_errors.append(prod_class)
                logging.warning('Unable to get value for {0} product class'.format(
                    prod_class))
            continue

        try:
            item_id = int(fields[1])
        except:
            item_id = 1

        if item_id in item_ids \
                or item_id > max_item_id \
                or last_item_id < item_id < 1000000:
            last_item_id = item_id = last_item_id + 1

        item_ids.add(item_id)

        try:
            item = MenuItem(id=item_id, name=name,
                            revenue_category=revenue_category,
                            product_class=product_class, sku=sku,
                            priceLvls=prices, byWeight=use_weight)
        except:
            logging.error(sys.exc_info()[0], sys.exc_info()[1])

        item_list.append(item)
        line = '{0},{1}'.format(update_type, item)
        if price_map:
            priced_items.append(line)
        else:
            unpriced_items.append(line)

    if priced_items:
        write_to_text_file(ig_priced_file, priced_items)
    if unpriced_items:
        write_to_text_file(ig_unpriced_file, unpriced_items)

    logging.info('IG import file creations complete.')
    messagebox.showinfo(
        title='Success',
        message='IG import files have been generated:\n\n{0}\n\n{1}\n\n'
                'Located in the following directory:\n\n{2}\n\n'
                'Please send these files to Agilysys'
                ' for importing into InfoGenesis'.format(
            os.path.basename(ig_priced_file),
            os.path.basename(ig_unpriced_file),
            os.path.dirname(ig_priced_file)
        )
    )


def get_revenue_categories(book):
    """Returns revenue categories dictionary from internal spreadsheet"""
    revenue_categories = dict()
    if isinstance(book, xlrd.book.Book):
        sheet = book.sheet_by_name('Revenue Categories')
        for row in range(1, sheet.nrows):
            try:
                revenue_categories[sheet.cell_value(row, 1)] = \
                    sheet.cell_value(row, 0)
            except IndexError:
                logging.error(
                    "oops, couldn't read row {0} from Revenue Categories"
                        .format(row))
    else:
        sheet = book['Revenue Categories']
        try:
            for row in sheet.rows:
                revenue_categories[row[1].value] = row[0].value
        except AttributeError:
            logging.error("oops, couldn't read row from Revenue Categories")

    return revenue_categories


def get_product_classes(book):
    """Returns product classes dict from internal spreadsheet"""
    product_classes = dict()
    if isinstance(book, xlrd.book.Book):
        sheet = book.sheet_by_name('Product Classes')
        for row in range(1, sheet.nrows):
            pc = sheet.cell_value(row, 1)
            pid = sheet.cell_value(row, 0)
            try:
                product_classes[pc] = pid
            except IndexError:
                logging.error(
                    "oops, couldn't read row {0} from Product Classes"
                        .format(row))
    else:
        sheet = book['Product Classes']
        for row in sheet.rows:
            try:
                product_classes[row[1].value] = row[0].value
            except:
                logging.error("oops, couldn't read row from Product Class")

    return product_classes


def get_file_type(filename):
    """Return the file type for filename"""
    file_extension = filename.rsplit('.', maxsplit=1)[1]
    if file_extension == 'xls' or file_extension == 'xlsx':
        logging.debug('Input file is xls, processing as EXCEL_FILE')
        return EXCEL_FILE
    elif file_extension == 'txt':
        file = codecs.open(filename, 'r', 'utf8')
        if len(file.readline().split(",")) > 20:
            return IG_EXPORT
        else:
            raise IOError('File is corrupt or contains incomplete data.')
    else:
        raise IOError('Unsupported file extension')


def convert_to_ig_format():
    """Initiates conversion from Excel spreadsheet to IG text file"""
    logging.debug('starting conversion to IG Format')
    options = {
        'title': 'Save As',
        'initialfile': os.path.join(os.path.dirname(in_file), 'MI_IMP.txt')
    }
    file_extension = in_file.rsplit('.', maxsplit=1)[1]
    if file_extension == 'xlsx':
        book = load_workbook(in_file, read_only=True)
        generate_standardized_ig_imports(book, in_file)
    elif file_extension == 'xls':
        book = open_workbook(in_file)
        sheet = book.sheet_by_index(0)

        if book.nsheets > 1:
            generate_standardized_ig_imports(book, in_file)
        else:
            file_save_path = save_file_as(options)
            if file_save_path:
                with codecs.open(file_save_path, 'w+', 'latin-1') as text_file:
                    if sheet.cell_value(1, 0) == 'modType':
                        generate_ig_import(book, text_file)
                    else:
                        generateIGUpdate(book, text_file)


def convert_to_excel():
    """Initiates conversion from IG Format to Excel spreadsheet"""
    file_parts = str(os.path.basename(in_file)).rsplit('.', maxsplit=1)
    default_filename = file_parts[0] + '_custom.xls'
    display_item_property_selections()
    root.wait_window(csWin)

    options = {'title': 'Save As',
               'initialfile': os.path.join(os.path.dirname(in_file),
                                           default_filename),
               'filetypes': file_type_filters}
    file_save_path = save_file_as(options)

    if file_save_path:
        logging.debug('saving output to {}'.format(file_save_path))
        generate_custom_excel_spreadsheet(file_save_path)


def build_ig_price_array(price_map):
    """Returns IG price array from dictionary of price levels."""
    prices = []
    price_is_negative = False
    for price_level, price in sorted(price_map.items()):
        if price != '':
            # Extract number from priceLvl
            if type(price_level) == str:
                level = str(price_level[price_level.find('l') + 1:])
            else:
                level = price_level
            price_sequence = ''
            if '(' in str(price) or ')' in str(price):
                price_is_negative = True
            try:
                price = '{0:.2f}'.format(float(str(price).strip('$(){}')))
            except ValueError:
                price = '0'
            if price_is_negative:
                price_sequence = '{0},(${1})'.format(level, price)
            else:
                price_sequence = '{0},${1}'.format(level, price)
            prices.append(price_sequence)

    record = '{' + ','.join(prices) + '}'
    return record


def duplicate_sku_finder(items):
    """Finds all items associated with duplicate barcodes"""
    logging.debug('duplicate_sku_finder started')
    duplicitous_items = []
    item_skus = []
    for item in items:
        sku = re.search(BARCODE_REGEX, item.sku)
        if sku:
            item_skus.append(sku.group(0))
    logging.debug('skus extracted from items')
    duplicate_skus = [x for x, y in collections.Counter(item_skus)
        .items() if y > 1]
    logging.debug('duplicate skus calculated')
    for item in items:
        sku = re.search(BARCODE_REGEX, item.sku)
        if sku and sku.group(0) in duplicate_skus:
            duplicitous_items.append(item)

    logging.debug('duplicate_sku_finder completed')
    return duplicitous_items


def duplicate_sku_test(barcode_file=None):
    """Debugging stub method to output duplicate barcodes"""
    global dupe_sku_check_enabled
    if not dupe_sku_check_enabled.get():
        logging.info('Barcode check disabled.  Skipping')
        return
    logging.debug('attempting duplicate sku test')
    if itemList:
        item_strings = []
        dupe_items = sorted(duplicate_sku_finder(itemList), key=lambda x: x.sku)
        for item in dupe_items:
            item_strings.append('"A",{}'.format(item))
    else:
        logging.error('Unable to complete duplicate_sku_test, itemList empty')
        return

    if item_strings:
        response = messagebox.askokcancel(
            title='Duplicate Barcodes Found',
            message='Would you like to save '
                    'duplicate barcodes to file?')
        if response:
            barcode_file = barcode_file or \
                           os.path.join(os.path.dirname(in_file),
                                        'item_duplicate_barcodes.txt')

            write_to_text_file(barcode_file, item_strings)
            logging.info('duplicate_barcodes output successfully')
            messagebox.showinfo(title='Success',
                                message='duplicate_barcodes'
                                        ' output successfully')
        else:
            logging.info('duplicate barcode output skipped')


def safeIntCast(value):
    """
    Attempts to cast value to an integer, falls back to a string if it fails.
    Will also set row_is_misaligned to True if int cast fails.
    """
    try:
        return str(int(value))
    except ValueError:
        global row_is_misaligned
        row_is_misaligned = True
        return str(value)


def hide_button(button):
    button.grid_remove()


def hide_all_buttons():
    for b in hideable_buttons:
        hide_button(b)


def show_button(button):
    button.grid()


def display_about():
    messagebox.showinfo(title='About', message=__version__)


def display_item_property_selections():
    """
    Display sub window for selecting properties to include in Excel spreadsheet
    """
    global csWin
    csWin = Toplevel(root)
    colSelectFrame = ttk.Frame(csWin)
    colSelectFrame.grid(column=0, row=1, sticky=(N, S, E, W))
    colSelectFrame.columnconfigure(0, weight=1)
    colSelectFrame.rowconfigure(1, weight=1)

    global checkbox_variable_map
    row = 5
    counter = 0

    for k, v in sorted(IG_FIELD_SEQUENCE.items(), key=lambda x: x[1]):
        col = 0
        if k not in checkbox_variable_map:
            checkbox_variable_map[k] = IntVar()
        if k != 'id' and k[:-3] != 'reserved':
            if counter % 2 == 0:
                col = 0
                row += 1
            else:
                col = 3
            l = ttk.Checkbutton(colSelectFrame, text=TEXT_HEADERS[k],
                                variable=checkbox_variable_map[k]).grid(
                column=col, row=row, sticky=(N, W))
            counter += 1

    ttk.Button(colSelectFrame, text='OK',
               command=csWin.destroy).grid(column=1, row=100)
    ttk.Button(colSelectFrame, text='Select All',
               command=select_all_properties).grid(column=1, row=0)
    return


def select_all_properties():
    """Selects all properties in checkVarMap"""
    global checkbox_variable_map
    global all_boxes_selected

    if all_boxes_selected:
        check_mark = 0
        all_boxes_selected = False
    else:
        check_mark = 1
        all_boxes_selected = True

    for k, v in MenuItem.IG_FIELD_SEQUENCE.items():
        if k not in checkbox_variable_map:
            checkbox_variable_map[k] = IntVar()
        # id is mandatory and shouldn't be included in variable map
        if k != 'id':
            checkbox_variable_map[k].set(check_mark)


def show_var_states(ttk_var):
    """prints state of ttk variables. For debugging purposes."""
    if type(ttk_var) is dict:
        for k, v in ttk_var.items():
            print('{0}: {1}'.format(k, v.get()))
    elif type(ttk_var) is list:
        for item in ttk_var:
            print('{0}: {1}'.format(item, item.get()))
    else:
        print('{0}: {1}'.format(ttk_var, ttk_var.get()))


def toggle_debug_logging(event=None):
    global log_level
    if not event:
        if debug_log_enabled.get():
            log_level = logging.DEBUG
        else:
            log_level = logging.WARNING
    elif log_level != logging.DEBUG:
        messagebox.showinfo(title='DEBUG', message='Debug logging enabled.')
        log_level = logging.DEBUG
        debug_log_enabled.set(True)
    else:
        return

    root_logger.setLevel(log_level)
    logging.error('Log Level set to {0}'.format(
        logging.getLevelName(log_level)))


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    else:
        return os.path.join(os.path.abspath("."), relative_path)


def main():
    global root, file_display_string, debug_log_enabled
    global FSOCK, simplifyButtons, hideable_buttons, button_ig
    global dupe_sku_check_enabled
    root = Tk()
    root.option_add('*tearOff', FALSE)
    root.title("Agilysys Import Export Tool")
    ICON = resource_path('resources/Format_Gears.ico')
    try:
        root.iconbitmap(default=ICON)
    except TclError:
        logging.error('Unable to locate icon at {0}'.format(ICON))

    file_display_string = StringVar()
    debug_log_enabled = BooleanVar()
    dupe_sku_check_enabled = BooleanVar()

    if log_level == logging.DEBUG:
        debug_log_enabled.set(True)
    else:
        debug_log_enabled.set(False)

    if not os.path.exists(APP_DIR):
        os.mkdir(APP_DIR)

    if getattr(sys, 'frozen', True):
        FSOCK = open(LOG_FILE, 'a+')
        sys.stderr = FSOCK

    menubar = Menu(root)
    menu_file = Menu(menubar)
    menu_debug_options = Menu(menubar)
    menu_help = Menu(menubar)
    menubar.add_cascade(menu=menu_file, label='File')
    menubar.add_cascade(menu=menu_help, label='Help')
    # Add Debug Menu only when not compiled
    if not getattr(sys, 'frozen', False):
        menubar.add_cascade(menu=menu_debug_options, label='Debug')

    menu_file.add_command(label='Open...', command=open_file)
    menu_file.add_command(label='Close', command=root.quit)

    menu_debug_options.add_command(label='Display Vars',
                                   command=lambda: show_var_states(
                                       checkbox_variable_map))

    menu_debug_options.add_checkbutton(
        label='Enable Debug Logging',
        variable=debug_log_enabled,
        command=toggle_debug_logging)

    menu_debug_options.add_checkbutton(label='Duplicate Barcode Check',
                                       variable=dupe_sku_check_enabled)

    menu_help.add_command(label='About', command=display_about)
    menu_help.add_command(
        label='Display Log',
        command=lambda: subprocess.call("explorer {}".format(
            os.path.dirname(LOG_FILE)), shell=True))

    mainframe = ttk.Frame(root, padding="3 3 12 12")
    mainframe.grid(column=0, row=1, sticky=(N, W, E, S))
    mainframe.columnconfigure(0, weight=1)
    mainframe.rowconfigure(1, weight=1)

    ttk.Label(mainframe, text="File:").grid(
        column=1, row=1, sticky=(N, W, E))
    openFile_entry = ttk.Entry(mainframe, width=40,
                               textvariable=file_display_string, state='disabled')
    openFile_entry.grid(column=1, row=2, sticky=(W, E))

    button_open = ttk.Button(mainframe, text='...', command=open_file)
    button_open.grid(column=2, row=2, sticky=(W))
    open_tooltip = ToolTip(button_open, 'Select the Excel file required to'
                                        ' generate the IG formatted CSV file')
    button_ig = ttk.Button(mainframe, text='Generate IG Update',
                           command=convert_to_ig_format)
    button_ig.grid(column=1, row=3)
    ig_tooltip = ToolTip(button_ig, 'Generate IG formatted CSV file(s)')

    button_excel = ttk.Button(mainframe, text='Create Excel File',
                              command=convert_to_excel)
    button_excel.grid(column=1, row=6)

    simplifyButtons = [button_excel]
    hideable_buttons = [button_excel, button_ig]

    for child in mainframe.winfo_children():
        child.grid_configure(padx=5, pady=5)

    root.bind_all("<Control-Alt-d>", toggle_debug_logging)
    root.config(menu=menubar)
    hide_all_buttons()
    root.mainloop()


if __name__ == "__main__":
    main()
